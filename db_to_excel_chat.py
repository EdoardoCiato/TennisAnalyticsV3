from __future__ import annotations

import os
import sqlite3

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# =========================
# CONFIG
# =========================
BASE_DIR       = os.path.dirname(os.path.abspath(__file__))
DB_PATH        = os.path.join(BASE_DIR, "tennis_abstract_new_version_merged.db")
AGG_V2_DB_PATH = os.path.join(BASE_DIR, "aggressiveness_v2.db")
XLSX_PATH      = os.path.join(BASE_DIR, "tennis_abstract_output.xlsx")

REPORT_TABLES = [
    "Serve Analysis",
    "Rally Analysis",
    "Attitude Analysis",
    "Tactics Analysis",
    "Efficiency Analysis",
    "Glossary",
]

POLARITY_MAP: dict[str, int] = {
    "1sts %": +1, "1st won %": +1, "2nd won %": +1, "Aces %": +1,
    "Serve Games hold %": +1, "Serve Pts won": +1, "Double Faults %": -1,
    "Unforced Errors": -1, "Rally Aggressiveness": +1, "Rally Winners / Unforced": +1,
    "Forehand Win %": +1, "Backhand Win %": +1, "Serve Stay/Match won": +1,
    "Return Aggressiveness": +1, "Break Consol. %": +1, "Tie Break won": +1,
    "Break Back %": +1, "Break Pts converted": +1, "Drop Freq.": +1,
    "Net Freq.": +1, "Serve & Volley Freq.": +1, "Crosscourt %": +1,
    "Down the Line %": +1, "Inside In %": +1, "Inside Out %": +1,
    "Ratio W/UE": +1, "Return Pts won": +1, "Winners": +1,
    "Break vs Break Pts": +1, "Games w/ Break Pts": +1,
}

AGG_POLARITY_MAP: dict[str, int] = {
    "coeff_serve":           +1,
    "coeff_rally":           +1,
    "coeff_attitude":        +1,
    "coeff_tactics":         +1,
    "coeff_efficiency":      +1,
    "coeff_global":          +1,
    "coeff_rally_length":    +1,
    "coeff_surface":         +1,
    "coeff_hard":            +1,
    "coeff_clay":            +1,
    "coeff_grass":           +1,
    "rank_aggressiveness":   -1,
    "pct_1-3_W%":            +1,
    "pct_4-6_W%":            +1,
    "pct_7-9_W%":            +1,
    "pct_10+_W%":            +1,
}

# Colonne per i 3 fogli aggressiveness
AGG_INDEX_COLS = [
    "player_name", "ranking_order",
    "coeff_serve", "coeff_rally", "coeff_attitude",
    "coeff_tactics", "coeff_efficiency",
    "coeff_rally_length", "coeff_surface",
    "coeff_global", "rank_aggressiveness",
]
SURFACE_COLS = [
    "player_name", "ranking_order",
    "coeff_hard", "coeff_clay", "coeff_grass",
    "Hard_M", "Clay_M", "Grass_M", "coeff_surface",
]
RALLY_LENGTH_COLS = [
    "player_name", "ranking_order",
    "1-3_W%", "4-6_W%", "7-9_W%", "10+_W%",
    "pct_1-3_W%", "pct_4-6_W%", "pct_7-9_W%", "pct_10+_W%",
    "coeff_rally_length",
]

# =========================
# STYLES
# =========================
FILL_GREEN      = PatternFill(start_color="00C853", end_color="00C853", fill_type="solid")
FILL_ORANGE     = PatternFill(start_color="FF9800", end_color="FF9800", fill_type="solid")
FILL_HDR_REPORT = PatternFill(start_color="1A3A6C", end_color="1A3A6C", fill_type="solid")
FILL_HDR_AGG    = PatternFill(start_color="2C5F8A", end_color="2C5F8A", fill_type="solid")
THIN   = Side(border_style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


# =========================
# HELPERS
# =========================
def list_tables(conn: sqlite3.Connection) -> list[str]:
    return [r[0] for r in conn.execute(
        "SELECT name FROM sqlite_master WHERE type='table' ORDER BY name"
    ).fetchall()]


def safe_sheet_name(name: str, used: set[str]) -> str:
    base, candidate, n = name[:31], name[:31], 1
    while candidate in used:
        suffix    = f"_{n}"
        candidate = base[:31 - len(suffix)] + suffix
        n        += 1
    used.add(candidate)
    return candidate


def to_float(value) -> float | None:
    if value is None:
        return None
    text = str(value).strip()
    if text in ("", "NA", "-", "None", "NULL"):
        return None
    try:
        return float(text.replace("%", "").replace(",", "."))
    except Exception:
        return None


def select_cols(df: pd.DataFrame, cols: list[str]) -> pd.DataFrame:
    return df[[c for c in cols if c in df.columns]]


# =========================
# SHEET STYLING
# =========================
def apply_header_style(ws, ncols: int, fill: PatternFill) -> None:
    for col in range(1, ncols + 1):
        cell           = ws.cell(row=1, column=col)
        cell.fill      = fill
        cell.font      = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = BORDER
    ws.row_dimensions[1].height = 36
    ws.freeze_panes             = "A2"
    ws.auto_filter.ref          = ws.dimensions


def apply_full_borders(ws, df: pd.DataFrame) -> None:
    for row in range(2, df.shape[0] + 2):
        for col in range(1, df.shape[1] + 1):
            cell           = ws.cell(row=row, column=col)
            cell.border    = BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center")


def apply_coloring(ws, df: pd.DataFrame, polarity_map: dict[str, int]) -> None:
    for col_idx, col_name in enumerate(df.columns, start=1):
        polarity = polarity_map.get(str(col_name))
        if polarity not in (+1, -1):
            continue
        nums  = [to_float(df.iloc[i][col_name]) for i in range(df.shape[0])]
        valid = [v for v in nums if v is not None]
        if len(valid) < 2:
            continue
        mn, mx = min(valid), max(valid)
        for row_idx, v in enumerate(nums, start=2):
            if v is None:
                continue
            cell = ws.cell(row=row_idx, column=col_idx)
            if polarity == +1:
                if v == mx: cell.fill = FILL_GREEN
                elif v == mn: cell.fill = FILL_ORANGE
            else:
                if v == mn: cell.fill = FILL_GREEN
                elif v == mx: cell.fill = FILL_ORANGE


def set_column_widths(ws, df: pd.DataFrame) -> None:
    for col_idx, col_name in enumerate(df.columns, start=1):
        max_len = len(str(col_name))
        for i in range(df.shape[0]):
            v = df.iloc[i, col_idx - 1]
            if v is not None:
                max_len = max(max_len, len(str(v)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 12), 30)


def write_sheet(
    writer,
    df: pd.DataFrame,
    sheet_name: str,
    used_sheets: set[str],
    header_fill: PatternFill,
    polarity_map: dict[str, int] | None = None,
) -> None:
    name = safe_sheet_name(sheet_name, used_sheets)
    df.to_excel(writer, sheet_name=name, index=False)
    ws = writer.book[name]
    apply_header_style(ws, df.shape[1], header_fill)
    apply_full_borders(ws, df)
    if polarity_map:
        apply_coloring(ws, df, polarity_map)
    set_column_widths(ws, df)
    print(f"   ✅ '{name}' ({df.shape[0]}x{df.shape[1]})")


# =========================
# MAIN
# =========================
def export_to_excel(db_path: str, agg_db_path: str, xlsx_path: str) -> None:
    print(f"📂 DB principale : {db_path}")
    print(f"📂 Agg v2 DB     : {agg_db_path}")
    print(f"📄 Output        : {xlsx_path}")

    used_sheets: set[str] = set()

    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:

        # ── Serve, Rally, Attitude, Tactics, Efficiency ──
        conn = sqlite3.connect(db_path)
        try:
            existing = set(list_tables(conn))
            tables   = [t for t in REPORT_TABLES if t in existing]
            missing  = [t for t in REPORT_TABLES if t not in existing]

            print(f"\n📊 Tabelle analisi ({len(tables)}):")
            for table in tables:
                df = pd.read_sql_query(f'SELECT * FROM "{table}"', conn)
                write_sheet(writer, df, table, used_sheets, FILL_HDR_REPORT, POLARITY_MAP)

            if missing:
                print(f"   ⚠️  Non trovate nel DB: {missing}")
        finally:
            conn.close()

        # ── Aggressiveness Index, Surface Index, Rally Length Index ──
        if not os.path.exists(agg_db_path):
            print(f"\n   ⚠️  {agg_db_path} non trovato")
            print("      Esegui prima: python .\\Agg_Coeff\\aggressiveness_v2.py")
        else:
            print(f"\n📊 Aggressiveness Indices (3 fogli):")
            conn = sqlite3.connect(agg_db_path)
            try:
                df_agg = pd.read_sql_query('SELECT * FROM "aggressiveness_index"', conn)

                write_sheet(writer,
                            select_cols(df_agg, AGG_INDEX_COLS),
                            "Aggressiveness Index",
                            used_sheets, FILL_HDR_AGG, AGG_POLARITY_MAP)

                write_sheet(writer,
                            select_cols(df_agg, SURFACE_COLS),
                            "Surface Index",
                            used_sheets, FILL_HDR_AGG, AGG_POLARITY_MAP)

                write_sheet(writer,
                            select_cols(df_agg, RALLY_LENGTH_COLS),
                            "Rally Length Index",
                            used_sheets, FILL_HDR_AGG, AGG_POLARITY_MAP)
            finally:
                conn.close()

    print(f"\n✅ Excel creato: {xlsx_path}  ({len(used_sheets)} fogli)")


if __name__ == "__main__":
    export_to_excel(DB_PATH, AGG_V2_DB_PATH, XLSX_PATH)